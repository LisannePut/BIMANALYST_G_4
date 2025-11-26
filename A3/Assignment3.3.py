import ifcopenshell
import ifcopenshell.geom
import os
import math
import numpy as np
import re

"""
BR18 Building Code Compliance Checker
======================================

This script analyzes IFC building models for compliance with BR18 evacuation route requirements.

BR18 Requirements Checked:
- Doors: clear opening width >= 800 mm (DOOR_MIN)
- Stairs: clear width >= 1000 mm (STAIR_MIN)
- Corridors: clear width >= 1300 mm (CORRIDOR_MIN) AND must link to a stair
- Stair flights: must be enclosed by walls (4-wall enclosure check)

Reference: BR18.pdf included in this folder
"""

# ============================================================================
# SECTION 1: IMPORTS AND CONFIGURATION
# ============================================================================

import os
import sys
import math
import re as _re
import numpy as np
import ifcopenshell
import ifcopenshell.geom

# Configuration Constants
# Configuration Constants
IFC_PATH = os.path.join(os.path.dirname(__file__), "model", "25-16-D-ARCH.ifc")
DOOR_MIN = 800  # Minimum door width in mm
STAIR_MIN = 1000  # Minimum stair width in mm
CORRIDOR_MIN = 1300  # Minimum corridor width in mm
BUFFER_BBOX = 1000.0  # Buffer for bounding box calculations
NEAREST_MAX = 30000.0  # Maximum distance for proximity checks

# Geometry settings for IFC shape extraction
GEOM_SETTINGS = ifcopenshell.geom.settings()
GEOM_SETTINGS.set(GEOM_SETTINGS.USE_WORLD_COORDS, True)


# ============================================================================
# SECTION 2: UTILITY FUNCTIONS - CACHING AND UNIT CONVERSION
# ============================================================================

# Performance optimization: Cache for storing computed bounding boxes
# Key: (entity_type, GlobalId), Value: (xmin, ymin, xmax, ymax) in mm
# This avoids expensive geometry recalculation when checking the same element multiple times
_BBOX_CACHE = {}


def to_mm(v):
    """Convert a dimension value to millimeters.
    """
    try:
        f = float(v)
    except Exception:
        return None
    return f if f > 100 else f * 1000.0


# ============================================================================
# SECTION 3: GEOMETRY EXTRACTION FUNCTIONS
# ============================================================================

def extract_dimensions_from_geometry(sp):
    """Extract width and length from IfcSpace geometry using 3D vertices.
    """
    try:
        verts = get_vertices(sp)
        if verts is not None and len(verts) > 0:
            # Convert from meters to millimeters (multiply by 1000)
            verts = verts * 1000.0
            minv = verts.min(axis=0)
            maxv = verts.max(axis=0)
            dims = maxv - minv
            # Return (longer dim, shorter dim) as (length, width)
            dim_sorted = sorted(dims[:2])  # Take X, Y (ignore Z height)
            if dim_sorted[1] > 0:  # Ensure width > 0
                return dim_sorted[1], dim_sorted[0]
    except Exception:
        pass
    
    return 0, 0



def get_vertices(product):
    """Extract 3D vertices (corner points) from an IFC product's geometry.
    
    This function attempts to create a 3D shape from the IFC element and
    extract all its vertex coordinates using world (absolute) coordinates.
    """
    try:
        shape = ifcopenshell.geom.create_shape(GEOM_SETTINGS, product)
        verts = np.array(shape.geometry.verts, dtype=float).reshape(-1, 3)
        return verts
    except (KeyboardInterrupt, Exception):
        # Return None for any geometry error (includes timeouts/interrupts)
        return None


# ============================================================================
# SECTION 4: PROPERTY EXTRACTION FUNCTIONS
# ============================================================================

# NOTE: get_bbox and get_door_midpoint were removed because the code
# now uses `get_vertices` + geometry-based centroids via
# `get_element_centroid`. They were unused and are deleted to keep
# the file clean.

def get_numeric(entity, names):
    """Extract a numeric property value from an IFC entity by searching multiple possible property names.
    """
    names_l = [n.lower() for n in names]
    
    # Step 1: Check direct attributes on the entity (e.g., entity.Width)
    for attr in dir(entity):
        try:
            if attr.lower() in names_l:
                v = getattr(entity, attr)
                r = to_mm(v)
                if r:
                    return r
        except Exception:
            continue
    
    # Step 2 & 3: Check property sets and quantity sets via IsDefinedBy relationships
    for rel in getattr(entity, 'IsDefinedBy', []) or []:
        try:
            if not rel.is_a('IfcRelDefinesByProperties'):
                continue
            pdef = rel.RelatingPropertyDefinition
            if pdef is None:
                continue
            if pdef.is_a('IfcPropertySet'):
                for p in getattr(pdef, 'HasProperties', []) or []:
                    try:
                        pname = (getattr(p, 'Name', '') or '').lower()
                        if any(n in pname for n in names_l):
                            if hasattr(p, 'NominalValue') and p.NominalValue is not None:
                                try:
                                    val = p.NominalValue.wrappedValue
                                except Exception:
                                    val = p.NominalValue
                                r = to_mm(val)
                                if r:
                                    return r
                    except Exception:
                        continue
            if pdef.is_a('IfcElementQuantity'):
                for q in getattr(pdef, 'Quantities', []) or []:
                    try:
                        qn = (getattr(q, 'Name', '') or '').lower()
                        if any(n in qn for n in names_l):
                            val = getattr(q, 'LengthValue', None) or getattr(q, 'AreaValue', None) or getattr(q, 'VolumeValue', None)
                            r = to_mm(val)
                            if r:
                                return r
                    except Exception:
                        continue
        except Exception:
            continue
    return None


def centroid_from_extruded(item):
    try:
        if not item or not item.is_a('IfcExtrudedAreaSolid'):
            return None
        pos = getattr(item, 'Position', None)
        loc = getattr(pos, 'Location', None) if pos else None
        coords = list(getattr(loc, 'Coordinates', [])) if loc else []
        x = float(coords[0]) if coords else 0.0
        y = float(coords[1]) if len(coords) > 1 else 0.0
        z = float(coords[2]) if len(coords) > 2 else 0.0
        # Convert location coords to mm (they come in meters or model units)
        x = x if x > 100 else x * 1000.0
        y = y if y > 100 else y * 1000.0
        z = z if z > 100 else z * 1000.0
        prof = getattr(item, 'SweptArea', None)
        if prof and prof.is_a('IfcRectangleProfileDef'):
            xd = float(getattr(prof, 'XDim', 0) or 0)
            yd = float(getattr(prof, 'YDim', 0) or 0)
            xd = xd if xd > 100 else xd * 1000.0
            yd = yd if yd > 100 else yd * 1000.0
            return (x + xd / 2.0, y + yd / 2.0, z + float(getattr(item, 'Height', 0)) / 2.0)
        return (x, y, z)
    except Exception:
        return None


def get_element_centroid(elem):
    """Get centroid using ifcopenshell.geom (same method as debug script)."""
    try:
        verts = get_vertices(elem)
        if verts is not None and len(verts) > 0:
            verts = verts * 1000.0  # Convert to mm
            return verts.mean(axis=0)
    except Exception:
        pass
    return None


# ============================================================================
# SECTION 5: SPACE CONNECTIVITY AND LINKAGE ANALYSIS
# ============================================================================

def build_space_bboxes(spaces):
    b = {}
    for sp in spaces:
        sid = getattr(sp, 'GlobalId', None) or str(id(sp))
        xmin = ymin = float('inf'); xmax = ymax = float('-inf')
        if getattr(sp, 'Representation', None):
            for rep in sp.Representation.Representations:
                for it in getattr(rep, 'Items', []) or []:
                    if it.is_a('IfcExtrudedAreaSolid'):
                        pos = getattr(it, 'Position', None)
                        loc = getattr(pos, 'Location', None) if pos else None
                        coords = list(getattr(loc, 'Coordinates', [])) if loc else []
                        x = float(coords[0]) if coords else 0.0
                        y = float(coords[1]) if len(coords) > 1 else 0.0
                        prof = getattr(it, 'SweptArea', None)
                        if prof and prof.is_a('IfcRectangleProfileDef'):
                            xd = float(getattr(prof, 'XDim', 0) or 0)
                            yd = float(getattr(prof, 'YDim', 0) or 0)
                            xd = xd if xd > 100 else xd * 1000.0
                            yd = yd if yd > 100 else yd * 1000.0
                            hx = xd / 2.0; hy = yd / 2.0
                            xmin = min(xmin, x - hx); ymin = min(ymin, y - hy)
                            xmax = max(xmax, x + hx); ymax = max(ymax, y + hy)
        b[sid] = (xmin, ymin, xmax, ymax) if xmin != float('inf') else None
    return b


def build_space_linkages(model, spaces):
    """Check if hallways connect to stair spaces via doors.
    Note:
        A hallway that connects to another hallway that connects to stairs is also linked.
    """
    # Identify stair and hallway spaces
    stair_spaces = {}
    hallway_spaces = {}

    spaces_list = list(spaces)
    for sp in spaces_list:
        sid = getattr(sp, 'GlobalId', None) or str(id(sp))
        name = (getattr(sp, 'Name', None) or '').lower()
        if 'stair' in name:
            stair_spaces[sid] = sp
        elif 'hallway' in name:
            hallway_spaces[sid] = sp

    # Build adjacency map between spaces (space_gid -> set(space_gid)) using doors
    adjacency = { (getattr(sp, 'GlobalId', None) or str(id(sp))): set() for sp in spaces_list }

    # Build helper map: opening_gid -> list of containing elements (walls etc.)
    opening_to_containers = {}
    for relv in model.by_type('IfcRelVoidsElement'):
        container = getattr(relv, 'RelatingBuildingElement', None)
        opening = getattr(relv, 'RelatedOpeningElement', None)
        if not opening:
            continue
        ogid = getattr(opening, 'GlobalId', None) or str(id(opening))
        if container is not None:
            opening_to_containers.setdefault(ogid, []).append(container)

    # We'll also record which door connects to which spaces and which containers its opening sits in
    door_map = {}
    door_container_map = {}

    for rel in model.by_type('IfcRelFillsElement'):
        opening = getattr(rel, 'RelatingOpeningElement', None)
        door = getattr(rel, 'RelatedBuildingElement', None)
        if not (opening and door and door.is_a('IfcDoor')):
            continue

        # Get opening centroid (try opening first, then door as fallback)
        oc = get_element_centroid(opening)
        if oc is None:
            oc = get_element_centroid(door)
        if oc is None:
            continue

        # Find all spaces that contain this opening
        connected_spaces = []
        margin = 500  # 500mm margin
        for sp in spaces_list:
            sp_gid = getattr(sp, 'GlobalId', None) or str(id(sp))
            verts = get_vertices(sp)
            if verts is not None and len(verts) > 0:
                verts = verts * 1000.0  # Convert to mm
                minv = verts.min(axis=0)
                maxv = verts.max(axis=0)
                if minv[0] - margin <= oc[0] <= maxv[0] + margin and \
                   minv[1] - margin <= oc[1] <= maxv[1] + margin:
                    connected_spaces.append(sp_gid)

        # Link all connected spaces pairwise in adjacency
        for i in range(len(connected_spaces)):
            for j in range(i + 1, len(connected_spaces)):
                a = connected_spaces[i]
                b = connected_spaces[j]
                adjacency.setdefault(a, set()).add(b)
                adjacency.setdefault(b, set()).add(a)

        # Record door -> spaces map
        dg = getattr(door, 'GlobalId', None) or str(id(door))
        door_map.setdefault(dg, set()).update(connected_spaces)

        # Record container types (walls etc.) for this opening so we can check compartmentation
        og = getattr(opening, 'GlobalId', None) or str(id(opening))
        containers = opening_to_containers.get(og, [])
        door_container_map[dg] = [c.is_a() for c in containers]

    # Now compute which hallways are linked to stairs.
    # Start from stairs and propagate through hallway nodes only.
    linked_hallways = set()
    from collections import deque
    q = deque()

    # Enqueue all hallways that are directly adjacent to a stair
    for stair_gid in stair_spaces:
        for nb in adjacency.get(stair_gid, set()):
            if nb in hallway_spaces and nb not in linked_hallways:
                linked_hallways.add(nb)
                q.append(nb)

    # BFS across hallway nodes only
    while q:
        current = q.popleft()
        for nb in adjacency.get(current, set()):
            if nb in hallway_spaces and nb not in linked_hallways:
                linked_hallways.add(nb)
                q.append(nb)

    # Prepare final map for all hallways
    space_linked_to_stairs = {}
    for sid in hallway_spaces:
        space_linked_to_stairs[sid] = (sid in linked_hallways)

    # Ensure all spaces have an entry (False for non-hallways)
    for sp in spaces_list:
        sid = getattr(sp, 'GlobalId', None) or str(id(sp))
        space_linked_to_stairs.setdefault(sid, False)

    return space_linked_to_stairs, door_map, door_container_map

def build_full_door_space_map(model, margin=1000):
    """Build a complete door->space connectivity map over ALL IfcSpace elements.
    """
    spaces_list = list(model.by_type('IfcSpace'))
    # Precompute space bboxes
    space_bboxes = {}
    for sp in spaces_list:
        sp_gid = getattr(sp, 'GlobalId', None) or str(id(sp))
        bb = _bbox2d_mm(sp)
        if bb:
            space_bboxes[sp_gid] = bb
    door_map_all = {}
    door_container_map_all = {}
    opening_to_containers = {}
    for relv in model.by_type('IfcRelVoidsElement'):
        container = getattr(relv, 'RelatingBuildingElement', None)
        opening = getattr(relv, 'RelatedOpeningElement', None)
        if not opening:
            continue
        ogid = getattr(opening, 'GlobalId', None) or str(id(opening))
        if container is not None:
            opening_to_containers.setdefault(ogid, []).append(container)
    for rel in model.by_type('IfcRelFillsElement'):
        opening = getattr(rel, 'RelatingOpeningElement', None)
        door = getattr(rel, 'RelatedBuildingElement', None)
        if not (opening and door and door.is_a('IfcDoor')):
            continue
        oc_open = get_element_centroid(opening)
        oc_door = get_element_centroid(door)
        oc = oc_open if oc_open is not None else oc_door
        if oc is None:
            continue
        dg = getattr(door, 'GlobalId', None) or str(id(door))
        connected_spaces = []
        # Centroid inclusion
        for sp_gid, (x1,y1,x2,y2) in space_bboxes.items():
            if (x1 - margin) <= oc[0] <= (x2 + margin) and (y1 - margin) <= oc[1] <= (y2 + margin):
                connected_spaces.append(sp_gid)
        # Door bbox intersection
        db = _bbox2d_mm(door)
        if db:
            dx1,dy1,dx2,dy2 = db
            db_exp = (dx1 - margin, dy1 - margin, dx2 + margin, dy2 + margin)
            for sp_gid, bb in space_bboxes.items():
                if sp_gid in connected_spaces:
                    continue
                if _bbox_intersect(db_exp, bb):
                    connected_spaces.append(sp_gid)
        # Opening bbox intersection (if available)
        ob = _bbox2d_mm(opening)
        if ob:
            ox1,oy1,ox2,oy2 = ob
            ob_exp = (ox1 - margin, oy1 - margin, ox2 + margin, oy2 + margin)
            for sp_gid, bb in space_bboxes.items():
                if sp_gid in connected_spaces:
                    continue
                if _bbox_intersect(ob_exp, bb):
                    connected_spaces.append(sp_gid)
        if connected_spaces:
            door_map_all.setdefault(dg, set()).update(connected_spaces)
        og = getattr(opening, 'GlobalId', None) or str(id(opening))
        containers = opening_to_containers.get(og, [])
        door_container_map_all[dg] = [c.is_a() for c in containers]
    # Also add mappings via space boundaries where the RelatedBuildingElement is a door
    try:
        for rb in model.by_type('IfcRelSpaceBoundary'):
            try:
                sp = getattr(rb, 'RelatingSpace', None)
                be = getattr(rb, 'RelatedBuildingElement', None)
                if not sp or not be or not getattr(be, 'is_a', lambda *_: False)('IfcDoor'):
                    continue
                sp_gid = getattr(sp, 'GlobalId', None) or str(id(sp))
                dg = getattr(be, 'GlobalId', None) or str(id(be))
                door_map_all.setdefault(dg, set()).add(sp_gid)
            except Exception:
                continue
    except Exception:
        pass
    return door_map_all, door_container_map_all


# ============================================================================
# SECTION 6: COMPLIANCE ANALYSIS FUNCTIONS - DOORS, STAIRS, CORRIDORS
# ============================================================================

def analyze_door(door, door_map, opening_map):
    """Analyze a door for BR18 compliance (minimum width requirement).
    """
    name = getattr(door, 'Name', None) or str(door)
    gid = getattr(door, 'GlobalId', None) or str(id(door))
    full = f"{name} [{gid}]"
    width = get_numeric(door, ['overallwidth', 'width', 'doorwidth'])
    op = opening_map.get(gid)
    if not width and op:
        if getattr(op, 'Representation', None):
            for rep in op.Representation.Representations:
                for it in getattr(rep, 'Items', []) or []:
                    if it.is_a('IfcExtrudedAreaSolid'):
                        prof = getattr(it, 'SweptArea', None)
                        if prof and prof.is_a('IfcRectangleProfileDef'):
                            w = to_mm(getattr(prof, 'YDim', None)) or to_mm(getattr(prof, 'XDim', None))
                            if w:
                                width = w
                                break
                if width:
                    break
    issues = []
    if width is None:
        issues.append('width unknown')
    elif width < DOOR_MIN:
        issues.append(f'width {width:.0f}mm < {DOOR_MIN}mm')
    linked = door_map.get(gid, set())
    return {'name': full, 'width_mm': width, 'linked_spaces': linked, 'issues': issues}


def analyze_stair(flight):
    """Analyze a stair flight for BR18 compliance (minimum width requirement).
    """
    name = getattr(flight, 'Name', None) or str(flight)
    gid = getattr(flight, 'GlobalId', None) or str(id(flight))
    full = f"{name} [{gid}]"
    width = get_numeric(flight, ['actual run width', 'actualrunwidth', 'run width', 'width', 'tread'])
    if width is None and getattr(flight, 'Representation', None):
        for rep in flight.Representation.Representations:
            for it in getattr(rep, 'Items', []) or []:
                if it.is_a('IfcExtrudedAreaSolid'):
                    prof = getattr(it, 'SweptArea', None)
                    if prof and prof.is_a('IfcRectangleProfileDef'):
                        xd = float(getattr(prof, 'XDim', 0) or 0)
                        yd = float(getattr(prof, 'YDim', 0) or 0)
                        xd = xd if xd > 100 else xd * 1000.0
                        yd = yd if yd > 100 else yd * 1000.0
                        width = max(xd, yd)
                        break
            if width is not None:
                break
    issues = []
    if width is None:
        issues.append('width unknown')
    elif width + 1e-6 < STAIR_MIN:
        issues.append(f'width {width:.0f}mm < {STAIR_MIN}mm')
    return {'name': full, 'width_mm': width, 'issues': issues}


# ============================================================================
# SECTION 7: STAIRCASE GROUPING AND ENCLOSURE ANALYSIS
# ============================================================================

def analyze_staircase_groups(model):
    """Group IfcStairFlight elements by their base staircase identifier extracted from the Name.

    Example flight names observed:
      Assembled Stair:Stair:1282665 Run 1
      Assembled Stair:Stair:1282665 Run 2
      Assembled Stair:Stair:1282665 Run 3

    We extract the numeric id after the last 'Stair:' token (here 1282665).
    Each unique id represents one staircase between two storeys according to user spec.

    Returns list of staircase dicts:
      { 'id': <numeric str>, 'flight_count': N, 'run_labels': [...], 'is_standard_3_run': bool }
    """
    flights = model.by_type('IfcStairFlight')
    groups = {}
    for fl in flights:
        name = (getattr(fl, 'Name', None) or '')
        # Extract last numeric sequence after 'Stair:'
        stair_id = None
        if 'Stair:' in name:
            parts = name.split('Stair:')
            # Take last part then isolate leading digits
            tail = parts[-1].strip()
            # tail may look like '1282665 Run 1' -> take digits at start
            import re as _re
            m = _re.match(r'(\d+)', tail)
            if m:
                stair_id = m.group(1)
        if not stair_id:
            continue
        g = groups.setdefault(stair_id, {'id': stair_id, 'flights': [], 'run_labels': []})
        g['flights'].append(fl)
        run_label = ''
        if 'Run' in name:
            # capture 'Run' part
            run_label = name.split('Run',1)[1].strip()
        g['run_labels'].append(run_label or 'unknown')
    # Build output list
    out = []
    for sid, g in groups.items():
        run_labels_norm = [rl for rl in g['run_labels']]
        # Determine if it matches expected 3-run pattern (Run 1, Run 2, Run 3)
        expected_set = {'1','2','3','Run 1','Run 2','Run 3'}
        # Simplify run label tokens
        simple_tokens = set()
        for rl in run_labels_norm:
            tok = rl.replace(':',' ').split()[0]
            simple_tokens.add(tok)
        is_standard = {'1','2','3'} <= simple_tokens or {'Run','1','2','3'} <= simple_tokens
        out.append({
            'id': sid,
            'flight_count': len(g['flights']),
            'run_labels': run_labels_norm,
            'is_standard_3_run': is_standard
        })
    return out


def analyze_staircase_group_enclosure(model, side_margin=300.0, wall_search_expand=500.0, debug_group_id=None):
    """Proximity enclosure check per staircase flight group.

    Improvement over previous version:
      - Prefer union of associated stair *space* bboxes (geometry-identified) instead of raw flight bboxes.
        This aligns group enclosure with space-level enclosure logic and avoids artificial open sides
        introduced by irregular flight arrangement (e.g. landing offsets making union larger than real shaft).
      - Fallback to flight union if no spaces found.
      - Optional debug output for a specific staircase id to inspect chosen bboxes and side coverage.

    Passing condition: all 4 sides covered by at least one wall bbox intersection.
    Returns list of dicts: {id, flight_count, sides_covered, missing_sides, has_issue, source}
    """
    groups = analyze_staircase_groups(model)
    if not groups:
        return []

    # Collect flights indexed by gid & names for quick membership
    flights = { (getattr(f,'GlobalId',None) or str(id(f))): f for f in model.by_type('IfcStairFlight') }

    # Geometry-based stair spaces mapping (space_gid -> {'space','name','flight_gids'})
    geom_stair_spaces = identify_stair_spaces_geometry(model)
    # Invert mapping flight_gid -> list(space_gid)
    flight_to_spaces = {}
    for sp_gid, rec in geom_stair_spaces.items():
        for fg in rec['flight_gids']:
            flight_to_spaces.setdefault(fg, set()).add(sp_gid)

    # Walls (standard + regular)
    walls = list(model.by_type('IfcWall')) + list(model.by_type('IfcWallStandardCase'))
    wall_bboxes = []
    for w in walls:
        wb = _bbox2d_mm(w)
        if wb:
            wall_bboxes.append(wb)

    results = []
    for g in groups:
        sid = g['id']
        group_flight_gids = []
        flight_bboxes = []
        for fl_gid, fl in flights.items():
            name = getattr(fl, 'Name', None) or ''
            if sid in name:  # flight belongs to this staircase id
                bb = _bbox2d_mm(fl)
                if bb:
                    flight_bboxes.append(bb)
                    group_flight_gids.append(fl_gid)
        if not flight_bboxes:
            results.append({'id': sid, 'flight_count': g['flight_count'], 'sides_covered': 0, 'missing_sides': ['left','right','bottom','top'], 'has_issue': True, 'source': 'none'} )
            continue

        # Attempt to derive union of associated stair spaces (those containing any group flight centroids)
        space_bboxes = []
        used_space_ids = set()
        for fg in group_flight_gids:
            for sp_gid in flight_to_spaces.get(fg, []):
                if sp_gid in used_space_ids:
                    continue
                sp_rec = geom_stair_spaces.get(sp_gid)
                if not sp_rec:
                    continue
                sp_bb = _bbox2d_mm(sp_rec['space'])
                if sp_bb:
                    space_bboxes.append(sp_bb)
                    used_space_ids.add(sp_gid)

        if space_bboxes:
            xs1 = min(b[0] for b in space_bboxes); ys1 = min(b[1] for b in space_bboxes)
            xs2 = max(b[2] for b in space_bboxes); ys2 = max(b[3] for b in space_bboxes)
            source = 'space_union'
        else:
            # Fallback: flight union (previous behavior)
            xs1 = min(b[0] for b in flight_bboxes); ys1 = min(b[1] for b in flight_bboxes)
            xs2 = max(b[2] for b in flight_bboxes); ys2 = max(b[3] for b in flight_bboxes)
            source = 'flight_union'

        # Build side strips (slightly shrink interior by side_margin/2 to reduce false missing side)
        # NOTE: side_margin kept; could be tuned if still missing sides erroneously.
        strips = {
            'left':   (xs1 - wall_search_expand, ys1 - wall_search_expand, xs1 + side_margin, ys2 + wall_search_expand),
            'right':  (xs2 - side_margin,       ys1 - wall_search_expand, xs2 + wall_search_expand, ys2 + wall_search_expand),
            'top':    (xs1 - wall_search_expand, ys2 - side_margin,       xs2 + wall_search_expand, ys2 + wall_search_expand),
        }
        covered = {k: False for k in strips}
        for wb in wall_bboxes:
            for k, strip in strips.items():
                if not covered[k] and _bbox_intersect(strip, wb):
                    covered[k] = True
            if all(covered.values()):
                break
        sides_covered = sum(1 for v in covered.values() if v)
        missing = [k for k, v in covered.items() if not v]
        has_issue = sides_covered < 3

        # Optional debug print for one target group id
        if debug_group_id and sid == str(debug_group_id):
            print(f"DEBUG StaircaseGroup {sid}: source={source} flights={len(group_flight_gids)} spaces={len(space_bboxes)} bbox=({xs1:.1f},{ys1:.1f},{xs2:.1f},{ys2:.1f}) sides_covered={sides_covered}/3 missing={missing}")
            if space_bboxes:
                for i, sb in enumerate(space_bboxes):
                    print(f"  DEBUG space_bbox[{i}]={sb}")
            for i, fb in enumerate(flight_bboxes[:5]):
                print(f"  DEBUG flight_bbox[{i}]={fb}")

        results.append({'id': sid, 'flight_count': g['flight_count'], 'sides_covered': sides_covered, 'missing_sides': missing, 'has_issue': has_issue, 'source': source})
    return results


def identify_stair_spaces_geometry(model):
    """Identify stair spaces by associating each IfcStairFlight centroid with a containing IfcSpace.

    This supplements name-based detection (spaces containing 'stair'). A space is flagged as a
    stair space if at least one stair flight centroid lies inside its 2D bbox (with margin).
    Returns a dict: {space_gid: {'space': space, 'name': name, 'flight_gids': set([...])}}
    """
    spaces = model.by_type('IfcSpace')
    flights = model.by_type('IfcStairFlight')
    # Precompute space bboxes
    space_bbox = {}
    for sp in spaces:
        gid = getattr(sp, 'GlobalId', None) or str(id(sp))
        bb = _bbox2d_mm(sp)
        if bb:
            space_bbox[gid] = bb
    # Get flight centroids
    flight_centroids = {}
    for fl in flights:
        verts = get_vertices(fl)
        if verts is not None and len(verts) > 0:
            verts = verts * 1000.0
            c = verts.mean(axis=0)
            flight_centroids[getattr(fl, 'GlobalId', None) or str(id(fl))] = (float(c[0]), float(c[1]))
    # Associate
    stair_spaces = {}
    margin = 300.0
    for fl_gid, (fx, fy) in flight_centroids.items():
        for sp_gid, bb in space_bbox.items():
            x1,y1,x2,y2 = bb
            if (x1 - margin) <= fx <= (x2 + margin) and (y1 - margin) <= fy <= (y2 + margin):
                sp = next((s for s in spaces if (getattr(s,'GlobalId',None) or str(id(s)))==sp_gid), None)
                if sp is None:
                    continue
                entry = stair_spaces.setdefault(sp_gid, {'space': sp, 'name': getattr(sp,'Name',None) or sp_gid, 'flight_gids': set()})
                entry['flight_gids'].add(fl_gid)
    # Merge name-based spaces even if no flight caught (keep original 5)
    for sp in spaces:
        name_l = (getattr(sp,'Name',None) or '').lower()
        if 'stair' in name_l:
            sp_gid = getattr(sp,'GlobalId',None) or str(id(sp))
            stair_spaces.setdefault(sp_gid, {'space': sp, 'name': getattr(sp,'Name',None) or sp_gid, 'flight_gids': set()})
    return stair_spaces


def _bbox2d_mm(entity):
    """Return (xmin,ymin,xmax,ymax) in mm for an entity using geometry verts; None on failure."""
    try:
        # Cache by type + GlobalId
        try:
            gid = getattr(entity, 'GlobalId', None) or str(id(entity))
            et = entity.is_a() if hasattr(entity, 'is_a') else type(entity).__name__
            key = (et, gid)
        except Exception:
            key = None
        if key and key in _BBOX_CACHE:
            return _BBOX_CACHE[key]

        verts = get_vertices(entity)
        if verts is None or len(verts) == 0:
            return None
        verts = verts * 1000.0
        minv = verts.min(axis=0)
        maxv = verts.max(axis=0)
        bb = (float(minv[0]), float(minv[1]), float(maxv[0]), float(maxv[1]))
        if key:
            _BBOX_CACHE[key] = bb
        return bb
    except Exception:
        return None


# ============================================================================
# SECTION 8: BOUNDING BOX AND GEOMETRIC HELPER FUNCTIONS
# ============================================================================

def _bbox_intersect(a, b, margin=0.0):
    ax1, ay1, ax2, ay2 = a
    bx1, by1, bx2, by2 = b
    return not (ax2 < bx1 - margin or bx2 < ax1 - margin or ay2 < by1 - margin or by2 < ay1 - margin)


# ============================================================================
# ============================================================================
# SECTION 9: STAIR FLIGHT ENCLOSURE & GEOMETRY HELPERS
# ============================================================================
# This section contains helper functions and the main 4-wall enclosure check
# for stair flights (analyze_stairflight_4wall_enclosure).
# ============================================================================

def identify_stair_spaces_geometry(model):
    """Identify stair spaces by associating each IfcStairFlight centroid with a containing IfcSpace.

    This supplements name-based detection (spaces containing 'stair'). A space is flagged as a
    stair space if at least one stair flight centroid lies inside its 2D bbox (with margin).
    Returns a dict: {space_gid: {'space': space, 'name': name, 'flight_gids': set([...])}}
    """
    spaces = model.by_type('IfcSpace')
    flights = model.by_type('IfcStairFlight')
    # Precompute space bboxes
    space_bbox = {}
    for sp in spaces:
        gid = getattr(sp, 'GlobalId', None) or str(id(sp))
        bb = _bbox2d_mm(sp)
        if bb:
            space_bbox[gid] = bb
    # Get flight centroids
    flight_centroids = {}
    for fl in flights:
        verts = get_vertices(fl)
        if verts is not None and len(verts) > 0:
            verts = verts * 1000.0
            c = verts.mean(axis=0)
            flight_centroids[getattr(fl, 'GlobalId', None) or str(id(fl))] = (float(c[0]), float(c[1]))
    # Associate
    stair_spaces = {}
    margin = 300.0
    for fl_gid, (fx, fy) in flight_centroids.items():
        for sp_gid, bb in space_bbox.items():
            x1,y1,x2,y2 = bb
            if (x1 - margin) <= fx <= (x2 + margin) and (y1 - margin) <= fy <= (y2 + margin):
                sp = next((s for s in spaces if (getattr(s,'GlobalId',None) or str(id(s)))==sp_gid), None)
                if sp is None:
                    continue
                entry = stair_spaces.setdefault(sp_gid, {'space': sp, 'name': getattr(sp,'Name',None) or sp_gid, 'flight_gids': set()})
                entry['flight_gids'].add(fl_gid)
    # Merge name-based spaces even if no flight caught (keep original 5)
    for sp in spaces:
        name_l = (getattr(sp,'Name',None) or '').lower()
        if 'stair' in name_l:
            sp_gid = getattr(sp,'GlobalId',None) or str(id(sp))
            stair_spaces.setdefault(sp_gid, {'space': sp, 'name': getattr(sp,'Name',None) or sp_gid, 'flight_gids': set()})
    return stair_spaces


def _bbox2d_mm(entity):
    """Return (xmin,ymin,xmax,ymax) in mm for an entity using geometry verts; None on failure."""
    try:
        # Cache by type + GlobalId
        try:
            gid = getattr(entity, 'GlobalId', None) or str(id(entity))
            et = entity.is_a() if hasattr(entity, 'is_a') else type(entity).__name__
            key = (et, gid)
        except Exception:
            key = None
        if key and key in _BBOX_CACHE:
            return _BBOX_CACHE[key]

        verts = get_vertices(entity)
        if verts is None or len(verts) == 0:
            return None
        verts = verts * 1000.0
        minv = verts.min(axis=0)
        maxv = verts.max(axis=0)
        bb = (float(minv[0]), float(minv[1]), float(maxv[0]), float(maxv[1]))
        if key:
            _BBOX_CACHE[key] = bb
        return bb
    except Exception:
        return None


def _bbox_intersect(a, b, margin=0.0):
    ax1, ay1, ax2, ay2 = a
    bx1, by1, bx2, by2 = b
    return not (ax2 < bx1 - margin or bx2 < ax1 - margin or ay2 < by1 - margin or by2 < ay1 - margin)




def analyze_stairflight_4wall_enclosure(model, side_margin=300.0, wall_search_expand=500.0):
    """Simple 4-wall enclosure check for IfcStairFlight entities.

    Returns list: {flight_name, flight_gid, fully_enclosed (bool), sides_covered, missing_sides}
    Debug and wall listing removed per user request.
    """
    flights = model.by_type('IfcStairFlight')
    if not flights:
        return []

    walls = list(model.by_type('IfcWall')) + list(model.by_type('IfcWallStandardCase'))
    wall_to_storey = {}
    flight_to_storey = {}
    
    for rel in model.by_type('IfcRelContainedInSpatialStructure'):
        parent = getattr(rel, 'RelatingStructure', None)
        if parent and parent.is_a('IfcBuildingStorey'):
            for e in getattr(rel, 'RelatedElements', []) or []:
                try:
                    gid = getattr(e, 'GlobalId', None) or str(id(e))
                    if e.is_a('IfcWall') or e.is_a('IfcWallStandardCase'):
                        wall_to_storey[gid] = parent
                    if e.is_a('IfcStairFlight'):
                        flight_to_storey[gid] = parent
                except Exception:
                    continue

    wall_bboxes_by_storey = {}
    results = []
    
    for flight in flights:
        flight_gid = getattr(flight, 'GlobalId', None) or str(id(flight))
        flight_name = getattr(flight, 'Name', None) or flight_gid
        
    # (Removed debug classification logic)
        
        fb = _bbox2d_mm(flight)
        
        if fb is None:
            results.append({
                'flight_name': flight_name,
                'flight_gid': flight_gid,
                'fully_enclosed': False,
                'sides_covered': 0,
                'missing_sides': ['left','right','top','bottom']
            })
            continue
        
        fx1, fy1, fx2, fy2 = fb
        storey = flight_to_storey.get(flight_gid)

        candidate_walls = []
        if storey:
            sid = getattr(storey, 'GlobalId', None) or str(id(storey))
            if sid not in wall_bboxes_by_storey:
                wall_bboxes_by_storey[sid] = []
                for w in walls:
                    w_gid = getattr(w, 'GlobalId', None) or str(id(w))
                    if wall_to_storey.get(w_gid) is storey:
                        wb = _bbox2d_mm(w)
                        if wb:
                            wall_bboxes_by_storey[sid].append((w_gid, wb))
            candidate_walls = wall_bboxes_by_storey[sid]
        
        # If no storey or no walls found for that storey, use all walls
        if not candidate_walls:
            if 'ALL' not in wall_bboxes_by_storey:
                wall_bboxes_by_storey['ALL'] = []
                for w in walls:
                    wb = _bbox2d_mm(w)
                    if wb:
                        wall_bboxes_by_storey['ALL'].append((getattr(w, 'GlobalId', None) or str(id(w)), wb))
            candidate_walls = wall_bboxes_by_storey['ALL']

        # Build all 4 side strips
        strips = {
            'left':   (fx1 - wall_search_expand, fy1 - wall_search_expand, fx1 + side_margin, fy2 + wall_search_expand),
            'right':  (fx2 - side_margin,       fy1 - wall_search_expand, fx2 + wall_search_expand, fy2 + wall_search_expand),
            'top':    (fx1 - wall_search_expand, fy2 - side_margin,       fx2 + wall_search_expand, fy2 + wall_search_expand),
            'bottom': (fx1 - wall_search_expand, fy1 - wall_search_expand, fx2 + wall_search_expand, fy1 + side_margin),
        }

        covered = {k: False for k in strips}
        for _, wb in candidate_walls:
            for k, strip in strips.items():
                if not covered[k] and _bbox_intersect(strip, wb):
                    covered[k] = True
        sides_covered = sum(1 for v in covered.values() if v)
        missing = [k for k, v in covered.items() if not v]
        fully_enclosed = (sides_covered == 4)
        results.append({
            'flight_name': flight_name,
            'flight_gid': flight_gid,
            'fully_enclosed': fully_enclosed,
            'sides_covered': sides_covered,
            'missing_sides': missing
        })

    return results


# ============================================================================
# SECTION 10: MAIN ANALYSIS FUNCTION
# ============================================================================

def main():
    """Main BR18 compliance analysis function - focused on corridor evacuation route checking.
    
    This "morning-style" optimized version analyzes only corridor/hallway spaces to reduce
    runtime. It skips detailed geometry analysis for non-corridor rooms.
    
    BR18 Requirements Checked:
    1. Doors: Clear opening width >= 800mm
    2. Stairs: Clear width >= 1000mm  
    3. Corridors: Clear width >= 1300mm AND must link to a stair (evacuation route)
    4. Stair flights: Must be enclosed (between two walls minimum)
    
    Process:
    1. Load IFC model and identify corridor + stair spaces by name
    2. Extract corridor dimensions from geometry (width/length)
    3. Build space connectivity graph using doors
    4. Check which corridors link to stairs (directly or via other corridors)
    5. Analyze all doors for width compliance
    6. Analyze all stair flights for width compliance
    7. Check stair flight enclosure (must be between at least 2 walls)
    8. Group flights into staircases and check group-level enclosure
    9. Generate timestamped Excel report with all findings
    
    Output:
    - Excel file: analysis_summary_YYYYMMDD_HHMMSS.xlsx in A3 folder
    - Console: Two lines with clickable file path
    
    Returns:
        None (writes Excel file as side effect)
    """
    model = ifcopenshell.open(IFC_PATH)
    all_spaces = model.by_type('IfcSpace')

    def _n(sp):
        """Helper function to get lowercase space name for token matching."""
        return (getattr(sp, 'Name', '') or '').lower()

    # Define tokens that identify corridor/hallway spaces
    hallway_tokens = ['hallway', 'corridor', 'passage', 'circulation']

    # Select corridor spaces only (these are the 18 we report on) + collect stair spaces for linkage graph
    corridor_spaces = [sp for sp in all_spaces if any(t in _n(sp) for t in hallway_tokens)]
    stair_spaces = [sp for sp in all_spaces if 'stair' in _n(sp)]

    # For building door/stair adjacency we include corridor + stair spaces only
    linkage_spaces = corridor_spaces + stair_spaces

    # Dictionary to store analysis results for each corridor
    # Key: space GlobalId, Value: dict with space details and analysis results
    analyses = {}
    
    for sp in corridor_spaces:  # Only analyse corridors
        sid = getattr(sp, 'GlobalId', None) or str(id(sp))
        # Try geometry first for accurate width
        length, width = extract_dimensions_from_geometry(sp)
        if width == 0:  # fallback to area/perimeter if geometry fails
            A = get_numeric(sp, ['area'])
            P = get_numeric(sp, ['perimeter'])
            if A and P:
                if A > 1000:
                    A_m2 = A / 1_000_000.0
                else:
                    A_m2 = A
                P_m = P if P > 100 else P * 1000.0
                try:
                    s = P_m / 2.0
                    w = (s - math.sqrt(max(0, s * s - 4 * A_m2))) / 2.0
                    length = (A_m2 / w) * 1000.0
                    width = w * 1000.0
                except Exception:
                    pass
        analyses[sid] = {
            'space': sp,
            'name': getattr(sp, 'Name', None),
            'type': getattr(sp, 'LongName', None),
            'width': width,
            'length': length,
        }

    # Build linkages (doors between corridor+stair subset)
    space_linked, door_map, door_container_map = build_space_linkages(model, linkage_spaces)

    for sid, a in analyses.items():
        a['links_to_stairs'] = space_linked.get(sid, False)
        a['is_elongated'] = (a['length'] >= 3 * a['width']) if a['width'] > 0 else False

    corridors = [(sid, a) for sid, a in analyses.items()]  # analyses already corridor-only
    # Build enhanced full door-space map (global scope) for richer stair entry detection
    door_map_all, door_container_map_all = build_full_door_space_map(model)
    doors = [analyze_door(d, door_map_all, door_container_map_all) for d in model.by_type('IfcDoor')]
    failing_doors = [d for d in doors if d['issues']]
    flights = model.by_type('IfcStairFlight')
    stairs = [analyze_stair(f) for f in flights]
    failing_stairs = [s for s in stairs if s['issues']]

    # Identify failing corridors (width < 1300mm OR no link to stairs)
    failing_corridors = []
    for sid, a in corridors:
        checks = 0; issues = []
        if a['width'] >= CORRIDOR_MIN:
            checks += 1
        else:
            issues.append(f"Width is {a['width']:.0f}mm")
        if a['links_to_stairs']:
            checks += 1
        else:
            issues.append(f"Does not link to stairs via doors/openings")
        if checks < 2:
            failing_corridors.append({'name': f"{a.get('name')} [{sid}]", 'issues': issues})

    # Determine passing corridors (those not in failing list)
    failing_sids = set()
    for fc in failing_corridors:
        n = fc.get('name','')
        if '[' in n and n.endswith(']'):
            sid_parsed = n.split('[')[-1][:-1]
            failing_sids.add(sid_parsed)

    passing_corridors = []
    for sid, a in corridors:
        if sid in failing_sids:
            continue
        checks_passed = []
        w = a.get('width', 0) or 0
        if w >= CORRIDOR_MIN:
            checks_passed.append('width')
        if a.get('links_to_stairs'):
            checks_passed.append('stairs')
        ratio = (a['length'] / a['width']) if a['width'] > 0 else 0
        passing_corridors.append({
            'name': f"{a.get('name')} [{sid}]",
            'width_mm': float(w),
            'length_mm': float(a.get('length', 0) or 0),
            'links_stairs': a.get('links_to_stairs', False),
            'ratio': float(ratio),
            'passed': checks_passed,
        })

    # ========================================================================
    # STAIR FLIGHT ENCLOSURE CHECKS
    # ========================================================================
    
    # Simple 4-wall enclosure check for IfcStairFlight entities (FAST - no space analysis needed)
    flight_4wall = analyze_stairflight_4wall_enclosure(model)

    # Staircase (flight group) summary - groups flights by staircase ID
    staircase_groups = analyze_staircase_groups(model)
    storey_count = len(model.by_type('IfcBuildingStorey'))
    expected_groups = max(storey_count - 2, 0) * 3 if storey_count >= 3 else max(storey_count - 1, 0) * 3

    # Staircase group proximity enclosure check
    group_enclosure = analyze_staircase_group_enclosure(model)
    failing_groups = [ge for ge in group_enclosure if ge['has_issue']]

    # Geometry-based stair space detection (may reveal additional stair spaces)
    geo_stair_spaces = identify_stair_spaces_geometry(model)

    # ========================================================================
    # EXCEL REPORT GENERATION
    # ========================================================================
    
    # Export summary to Excel (.xlsx) only with timestamp
    import os
    from datetime import datetime
    base_dir = os.path.dirname(__file__)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    xlsx_path = os.path.join(base_dir, f'analysis_summary_{timestamp}.xlsx')

    def _write_xlsx(path):
        """Generate Excel report with BR18 compliance results.
        
        Creates a single-sheet Excel file with:
        - Requirements section (rows 1-6): Lists BR18 rules being checked
        - Table header (row 7): Column headers with bold gray background
        - Data rows (8-11): One row per category (Doors, Corridors, Stairs, Stair flights)
        
        Table structure:
        Column A: Category name
        Column B: Passing count (elements that meet requirements)
        Column C: Failing count (elements that violate requirements)
        Column D: Failing element IDs (vertical list, newline-separated)
        Column E: Reason for failure (vertical list, newline-separated)
        
        Formatting:
        - Text wrapping enabled for columns D & E (multi-line content)
        - Fixed column widths for readability
        - Vertical alignment = top for better readability of lists
        
        Args:
            path: Full file path where Excel file will be saved
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        wb = Workbook()

        # Single sheet matching requested format
        ws = wb.active
        ws.title = 'IFC_Compliance_Report'

        # Requirements section at top (rows 1-6)
        ws.append(['Requirements'])
        ws['A1'].font = Font(bold=True, size=12)
        ws.append(['- Doors: clear opening width ≥ 800 mm'])
        ws.append(['- Corridors: clear width ≥ 1300 mm AND must link to a stair via a door/opening'])
        ws.append(['- Stairs: clear flight width ≥ 1000 mm'])
        ws.append(['- Stair flights: must be enclosed by 4 walls (left, right, top, bottom)'])
        ws.append([''])

        # Table header (row 7) with formatting
        ws.append(['Category', 'Passing count', 'Failing count', "Failing element ID's", 'Reason for failure'])
        for c in ('A','B','C','D','E'):
            ws[f"{c}7"].font = Font(bold=True)
            ws[f"{c}7"].fill = PatternFill(start_color='FFEFEFEF', end_color='FFEFEFEF', fill_type='solid')
            ws[f"{c}7"].alignment = Alignment(horizontal='center')

        # Row 8: Doors compliance data
        door_fail_ids = []  # we don't have IDs directly for failing doors in summary; leave empty or collect if available
        door_reasons = []
        for d in failing_doors:
            door_fail_ids.append('')  # ID not stored; could be parsed from d['name'] if needed
            door_reasons.append('; '.join(d.get('issues', [])))
        ws.append([
            'Doors',
            (len(doors) - len(failing_doors)),  # Passing count
            len(failing_doors),  # Failing count
            '\n'.join(door_fail_ids) if door_fail_ids else '',  # Failing IDs (vertical list)
            '; '.join(door_reasons) if door_reasons else ''  # Reasons (semicolon-separated)
        ])
        # Enable text wrapping for IDs and reasons columns
        ws['D8'].alignment = Alignment(wrap_text=True, vertical='top')
        ws['E8'].alignment = Alignment(wrap_text=True, vertical='top')

        # Row 9: Corridors compliance data
        corridor_fail_ids = []
        corridor_reasons = []
        for c in failing_corridors:
            # Extract element ID from name like "Hallway:XXXXX [GID]" if present
            nm = c.get('name','')
            try:
                if ':' in nm:
                    # Parse ID from format "Hallway:1234567 [...]"
                    corridor_fail_ids.append(nm.split(':',1)[1].split()[0])
                else:
                    corridor_fail_ids.append('')
            except Exception:
                corridor_fail_ids.append('')
            corridor_reasons.append('; '.join(c.get('issues', [])))
        ws.append([
            'Corridors',
            len(passing_corridors),  # Passing count
            len(failing_corridors),  # Failing count
            '\n'.join(corridor_fail_ids) if corridor_fail_ids else '',  # Failing IDs (vertical list)
            '\n'.join(corridor_reasons) if corridor_reasons else ''  # Reasons (vertical list with newlines)
        ])
        ws['D9'].alignment = Alignment(wrap_text=True, vertical='top')
        ws['E9'].alignment = Alignment(wrap_text=True, vertical='top')

        # Row 10: Stairs (width) compliance data
        stair_fail_ids = []
        stair_reasons = []
        for s in failing_stairs:
            # Stair ID not parsed; leave blank or parse from name if pattern exists
            stair_fail_ids.append('')
            stair_reasons.append('; '.join(s.get('issues', [])))
        ws.append([
            'Stairs (width)',
            (len(stairs) - len(failing_stairs)),
            len(failing_stairs),
            '\n'.join(stair_fail_ids) if stair_fail_ids else '',
            '; '.join(stair_reasons) if stair_reasons else ''
        ])
        ws['D10'].alignment = Alignment(wrap_text=True, vertical='top')
        ws['E10'].alignment = Alignment(wrap_text=True, vertical='top')

        # Row 11: Stair flights enclosure compliance data
        failing_flights = [f for f in flight_4wall if not f.get('fully_enclosed')]
        passing_flights = [f for f in flight_4wall if f.get('fully_enclosed')]
        flight_fail_ids = []
        flight_reasons = []
        for f in failing_flights:
            # Extract staircase ID and Run from name like "Assembled Stair:Stair:1282665 Run 3"
            # Format as "1282665 Run 3" for readability
            name = f.get('flight_name', '')
            try:
                if 'Stair:' in name and 'Run' in name:
                    # Extract number after last 'Stair:'
                    stair_part = name.split('Stair:')[-1]
                    stair_id = stair_part.split()[0]  # Get numeric ID like "1282665"
                    # Extract Run number
                    run_part = name.split('Run', 1)[1].strip() if 'Run' in name else ''
                    flight_fail_ids.append(f"{stair_id} Run {run_part}")
                else:
                    flight_fail_ids.append(name)
            except Exception:
                flight_fail_ids.append(name)
            # Show how many sides are covered (e.g., "sides_covered=2/4")
            flight_reasons.append(f"sides_covered={f.get('sides_covered',0)}/4")
        ws.append([
            'Stair flights (4-wall enclosure)',
            len(passing_flights),  # Passing count
            len(failing_flights),  # Failing count
            '\n'.join(flight_fail_ids) if flight_fail_ids else '',  # Failing IDs (vertical list)
            '\n'.join(flight_reasons) if flight_reasons else ''  # Reasons (vertical list)
        ])
        ws['D11'].alignment = Alignment(wrap_text=True, vertical='top')
        ws['E11'].alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-size columns for optimal readability
        widths = {'A': 32, 'B': 16, 'C': 16, 'D': 36, 'E': 48}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        wb.save(path)

    # Generate Excel file with timestamp
    try:
        import openpyxl  # noqa: F401
        _write_xlsx(xlsx_path)
        
        # Print minimal console output with clickable file link (OSC 8 hyperlink protocol)
        # This creates terminal hyperlinks that work in VS Code, iTerm2, and other modern terminals
        import urllib.parse as _url
        xlsx_url = f"file://{_url.quote(xlsx_path)}"
        
        def _osc8_link(url: str, text: str):
            """Create OSC 8 terminal hyperlink (clickable link in terminal).
            
            Format: ESC ]8;;<URL> ESC \\ <TEXT> ESC ]8;; ESC \\
            Where ESC = \u001b and ESC \\ = \u0007
            """
            return f"\u001b]8;;{url}\u0007{text}\u001b]8;;\u0007"
        
        # Output: Two lines only
        # Line 1: Message + clickable file path
        # Line 2: Explicit "Click:" instruction + clickable link text
        print("Results of the evacuation check:", _osc8_link(xlsx_url, xlsx_path))
        print("Click:", _osc8_link(xlsx_url, "Open Excel (.xlsx)"))
    except Exception as e:
        print(f"Error: Could not generate Excel file. {e}")


if __name__ == '__main__':
    main()
